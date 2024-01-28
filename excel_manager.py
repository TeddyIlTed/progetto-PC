from openpyxl import load_workbook, Workbook
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import simpledialog


class ExcelManager:
    def __init__(self):
        self.config_file = 'config.txt'
        self.file_path = self.read_config()

        # Definizione dell'elenco dei mezzi
        self.mezzi = [
            "01 - Hyundai H1", "02 - Dacia Sandero", "03 - Ford Ranger",
            "04 - Ford Transit (telonato)", "05 - Ford Transit (9 posti)",
            "06 - Camion con gru", "07 - Camion maxi emergenza",
            "08 - Toyota Hilux", "09 - Fiat Ducato", "010 - Cestello/PLE",
            "011 - Quad"
        ]

        if not self.file_path:
            self.ask_save_location()
        else:
            self.load_or_create_workbook()

    def read_config(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r') as file:
                return file.read().strip()
        return None

    def write_config(self, path):
        with open(self.config_file, 'w') as file:
            file.write(path)

    def ask_save_location(self):
        root = tk.Tk()
        root.withdraw()
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Salva come"
        )
        root.destroy()

        if save_path:
            self.file_path = save_path
            self.write_config(self.file_path)
            self.load_or_create_workbook()
        else:
            raise Exception("È necessario scegliere un percorso per salvare il file Excel.")

    def load_or_create_workbook(self):
        if not os.path.exists(self.file_path):
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)
            self.create_sheets()
            self.workbook.save(self.file_path)
        else:
            self.workbook = load_workbook(self.file_path)

    def create_sheets(self):
        self.add_headers(self.workbook.create_sheet("Generale"), include_mezzo=True)
        for mezzo in self.mezzi:
            sanitized_name = self.sanitize_sheet_name(mezzo)
            self.add_headers(self.workbook.create_sheet(sanitized_name), include_mezzo=False)

    def add_headers(self, sheet, include_mezzo):
        headers = ["Numero servizio", "Data", "Orario uscita", "Orario rientro",
                   "Km iniziali", "Km finali", "Matricola autista", "Matricola CS",
                   *["Matricola volontario " + str(i) for i in range(1, 9)],
                   "Litri rifornimento", "Motivo uscita"]
        if include_mezzo:
            headers.insert(2, "Mezzo")
        sheet.append(headers)

    def sanitize_sheet_name(self, name):
        return name.replace("/", "-")
    
    def get_initial_km(self, mezzo):
        sheet = self.workbook['Generale']
        for row in reversed(list(sheet.iter_rows(values_only=True))):
            if row[2] == mezzo and row[6]:  # Controlla se 'Mezzo' e 'Km finali' sono presenti
                return row[6]  # Restituisce 'Km finali' dell'ultimo servizio come 'Km iniziali'
        # Se non ci sono servizi precedenti o i 'Km finali' non sono presenti, chiedi i km iniziali
        return self.ask_initial_km(mezzo)

    def get_last_km(self, mezzo):
        sheet = self.workbook['Generale']
        for row in reversed(list(sheet.iter_rows(values_only=True))):
            if row[2] == mezzo:  # Colonna 'Mezzo'
                return row[6]  # Colonna 'Km finali' dell'ultimo servizio
        return None


    def ask_initial_km(self, mezzo):
        initial_km = simpledialog.askinteger("Km iniziali", f"Inserisci i km iniziali per il mezzo {mezzo}:")
        return str(initial_km) if initial_km is not None else "0"

    def add_data_to_sheet(self, data):
        mezzo = data["Mezzo"]
        identifier = (data["Data (formato DD/MM/YY)"], data["Orario uscita (formato HH:MM)"], mezzo)
        
        # Imposta i km iniziali in base all'ultimo servizio del mezzo
        initial_km = self.get_last_km(mezzo)
        data["Km iniziali"] = initial_km if initial_km is not None else "0"

        # Aggiornamento dei fogli
        if not self.check_for_duplicates('Generale', identifier, include_mezzo=True):
            self.update_sheet('Generale', data, include_mezzo=True)
        
        if not self.check_for_duplicates(mezzo, identifier, include_mezzo=False):
            self.update_sheet(mezzo, data, include_mezzo=False)
        
        self.workbook.save(self.file_path)

    def check_for_duplicates(self, sheet_name, identifier, include_mezzo):
        sheet = self.workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_identifier = (
                row[1],  # Data
                row[3 if include_mezzo else 2],  # Orario uscita
                row[2] if include_mezzo else sheet_name  # Mezzo
            )
            if row_identifier == identifier:
                return True
        return False

    def update_sheet(self, sheet_name, data, include_mezzo):
            sheet = self.workbook[sheet_name]
            # Aggiungi i dati nel foglio
            row_data = [data["Numero servizio"], data["Data (formato DD/MM/YY)"]]
            if include_mezzo:
                row_data.append(data["Mezzo"])
            row_data.extend([
                data["Orario uscita (formato HH:MM)"], data["Orario rientro (formato HH:MM)"],
                self.get_initial_km(data["Mezzo"]), data["Km finali"],
                data["Matricola autista"], data["Matricola CS"]
            ])
            row_data.extend([data[f"Matricola volontario {i}"] for i in range(1, 9)])
            row_data.extend([data["Litri rifornimento"], data["Motivo uscita"]])
            sheet.append(row_data)

    def set_file_path(self, file_path):
        self.file_path = file_path
        self.load_or_create_workbook()